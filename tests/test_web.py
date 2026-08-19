import re
from io import BytesIO
from datetime import date, datetime, time, timedelta

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app import app
from database.database import Base, get_db
from database.models import AccessRecord, Administrator, AuthorizedPersonnel, DailyOperation, Student
from security import hash_password


def test_scanner_page_loads():
    with TestClient(app) as client:
        response = client.get("/")
        assert response.status_code == 200
        assert "ESPERANDO" in response.text
        assert "ESCÁNER" in response.text and "SCHOOL" in response.text


def test_public_stats_api():
    with TestClient(app) as client:
        response = client.get("/api/stats/today")
        assert response.status_code == 200
        assert {"entries", "exits", "denied", "inside"}.issubset(response.json())


def test_admin_trash_and_day_detail_pages_render():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    historical_date = date.today() - timedelta(days=1)
    with Session(engine) as db:
        db.add(Administrator(username="admin-test", password_hash=hash_password("Prueba1234"),
                             nombre="Admin Test", rol="ADMIN", activo=True))
        student = Student(student_id="SAU-WEB-001", matricula="WEB001", nombres="Grace",
                          apellido_paterno="Hopper", apellido_materno="", carrera="Computación",
                          plantel="Campus Test", turno="Matutino", activo=True,
                          fecha_vencimiento=date.today() + timedelta(days=30))
        db.add(student)
        db.flush()
        operation = DailyOperation(fecha=historical_date, estado="CERRADA",
                                   abierta_at=datetime.combine(historical_date, time(8)),
                                   cerrada_at=datetime.combine(historical_date, time(18)))
        db.add(operation)
        person = AuthorizedPersonnel(personnel_id="SAU-PER-WEB-001", numero_empleado="PER001",
            nombres="Personal", apellido_paterno="Prueba", apellido_materno="",
            tipo_personal="DOCENTE", area="Ingeniería", plantel="Campus Test", turno="Matutino",
            activo=True, fecha_vencimiento=date.today() + timedelta(days=30))
        db.add(person)
        db.add(AccessRecord(student_id=student.id, tipo_movimiento="ENTRADA", fecha=historical_date,
                            hora=time(8, 30), datetime=datetime.combine(historical_date, time(8, 30)),
                            resultado="AUTORIZADO", motivo="ACCESO AUTORIZADO", dispositivo="TEST"))
        db.commit()
        student_id = student.id
        operation_id = operation.id
        record_id = db.query(AccessRecord).one().id
        person_id = person.id

    def override_db():
        with Session(engine) as db:
            yield db

    app.dependency_overrides[get_db] = override_db
    try:
        with TestClient(app) as client:
            login_page = client.get("/admin/login")
            token = re.search(r'name="csrf_token" value="([^"]+)"', login_page.text).group(1)
            login = client.post("/admin/login", data={"username": "admin-test", "password": "Prueba1234",
                                                      "csrf_token": token}, follow_redirects=False)
            assert login.status_code == 303
            students = client.get("/admin/students")
            records = client.get("/admin/records")
            days = client.get("/admin/days")
            detail = client.get(f"/admin/days/{operation_id}")
            personnel = client.get("/admin/personnel")
            inside = client.get("/admin/inside")
            assert all(response.status_code == 200 for response in (students, records, days, detail, personnel, inside))
            assert "Papelera" in students.text and "Papelera" in records.text and "Papelera" in days.text
            assert all("Seleccionar" in response.text for response in (students, records, days))
            assert "Movimientos de la jornada" in detail.text and "WEB001" in detail.text
            assert "PER001" in personnel.text and "Seleccionar todos" in personnel.text
            assert "Docentes y personal autorizado" in inside.text
            dashboard = client.get("/admin")
            assert "movement-period" in dashboard.text and "career-select" in dashboard.text
            yearly_chart = client.get("/admin/charts/data", params={"period": "year", "career": "Computación"})
            daily_chart = client.get("/admin/charts/data", params={"period": "day"})
            assert yearly_chart.status_code == 200 and len(yearly_chart.json()["labels"]) == 12
            assert yearly_chart.json()["selected_career"] == "Computación"
            assert len(daily_chart.json()["labels"]) == 24
            assert "PERSONAS QUE ENTRARON" in records.text
            assert "Consultas rápidas" in records.text and "Personas sin salida" in records.text
            report_params = {"from": historical_date.isoformat(), "to": historical_date.isoformat(),
                             "carrera": "Computación"}
            excel_report = client.get("/admin/records/entries/report/xlsx", params=report_params)
            pdf_report = client.get("/admin/records/entries/report/pdf", params=report_params)
            csv_report = client.get("/admin/records/entries/report/csv", params=report_params)
            assert excel_report.status_code == 200 and excel_report.content[:2] == b"PK"
            assert pdf_report.status_code == 200 and pdf_report.content.startswith(b"%PDF")
            assert "WEB001" in csv_report.content.decode("utf-8-sig")
            workbook = load_workbook(BytesIO(excel_report.content), read_only=True)
            assert workbook.sheetnames == ["Resumen", "Quiénes entraron", "Personas únicas"]
            assert workbook["Resumen"]["B6"].value == 1
            assert workbook["Quiénes entraron"]["C2"].value == "WEB001"
            assert workbook["Personas únicas"]["A2"].value == "WEB001"
            student_credential = client.get(f"/admin/students/{student_id}/credential")
            personnel_credential = client.get(f"/admin/personnel/{person_id}/credential")
            assert student_credential.status_code == 200 and student_credential.content.startswith(b"%PDF")
            assert personnel_credential.status_code == 200 and personnel_credential.content.startswith(b"%PDF")
            assert "Credencial-WEB001.pdf" in student_credential.headers["content-disposition"]
            admin_token = re.search(r'name="csrf_token" value="([^"]+)"', students.text).group(1)
            assert client.post("/admin/students/bulk", data={"csrf_token": admin_token,
                               "action": "delete", "ids": str(student_id)},
                               follow_redirects=False).status_code == 303
            assert client.post("/admin/records/bulk", data={"csrf_token": admin_token,
                               "action": "delete", "ids": str(record_id)},
                               follow_redirects=False).status_code == 303
            assert client.post("/admin/days/bulk", data={"csrf_token": admin_token,
                               "action": "delete", "ids": str(operation_id)},
                               follow_redirects=False).status_code == 303
            assert client.post("/admin/personnel/bulk", data={"csrf_token": admin_token,
                               "action": "delete", "ids": str(person_id)}, follow_redirects=False).status_code == 303
            assert "WEB001" in client.get("/admin/students?papelera=1").text
            assert "ACCESO AUTORIZADO" in client.get("/admin/records?papelera=1").text
            assert historical_date.strftime("%d/%m/%Y") in client.get("/admin/days?papelera=1").text
            assert "PER001" in client.get("/admin/personnel?papelera=1").text
            client.post(f"/admin/students/{student_id}/restore", data={"csrf_token": admin_token})
            client.post(f"/admin/records/{record_id}/restore", data={"csrf_token": admin_token})
            client.post(f"/admin/days/{operation_id}/restore", data={"csrf_token": admin_token})
            client.post("/admin/personnel/bulk", data={"csrf_token": admin_token,
                        "action": "restore", "ids": str(person_id)})
            with Session(engine) as db:
                assert db.get(Student, student_id).eliminado is False
                assert db.get(AccessRecord, record_id).eliminado is False
                assert db.get(DailyOperation, operation_id).eliminado is False
                assert db.get(AuthorizedPersonnel, person_id).eliminado is False
    finally:
        app.dependency_overrides.clear()
