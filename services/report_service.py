from __future__ import annotations

import csv
import io
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config import settings
from database.models import AccessRecord


HEADERS = ["Fecha", "Hora", "Identificador", "Persona", "Grupo o área", "Movimiento", "Resultado", "Motivo"]
ENTRY_HEADERS = ["Fecha", "Hora", "Identificador", "Persona", "Tipo", "Carrera o área", "Grupo", "Plantel"]


def _row(record: AccessRecord) -> list[str]:
    person = record.person
    return [
        record.datetime.strftime("%d/%m/%Y"),
        record.datetime.strftime("%H:%M:%S"),
        record.student.matricula if record.student else (record.personnel.numero_empleado if record.personnel else "—"),
        person.nombre_completo if person else "No identificado",
        record.student.carrera if record.student else (record.personnel.area if record.personnel else "—"),
        record.tipo_movimiento or "—",
        record.resultado,
        record.motivo,
    ]


def create_csv(records: list[AccessRecord]) -> bytes:
    stream = io.StringIO(newline="")
    stream.write("\ufeff")
    writer = csv.writer(stream)
    writer.writerow(HEADERS)
    writer.writerows(_row(record) for record in records)
    return stream.getvalue().encode("utf-8")


def create_excel(records: list[AccessRecord]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimientos"
    sheet.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="173C35")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for record in records:
        sheet.append(_row(record))
    widths = [13, 12, 16, 34, 30, 16, 16, 32]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_pdf(records: list[AccessRecord]) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    entries = sum(1 for record in records if record.tipo_movimiento == "ENTRADA" and record.resultado == "AUTORIZADO")
    exits = sum(1 for record in records if record.tipo_movimiento == "SALIDA" and record.resultado == "AUTORIZADO")
    denied = sum(1 for record in records if record.resultado == "DENEGADO")
    story = [
        Paragraph(settings.INSTITUTION_NAME, styles["Title"]),
        Paragraph("Reporte de movimientos de acceso", styles["Heading2"]),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · Entradas: {entries} · Salidas: {exits} · Rechazados: {denied}", styles["BodyText"]),
        Spacer(1, 8 * mm),
    ]
    data = [HEADERS] + [_row(record) for record in records]
    table = Table(data, repeatRows=1, colWidths=[22*mm, 20*mm, 25*mm, 43*mm, 43*mm, 25*mm, 25*mm, 42*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173C35")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DDD9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F5")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(table)
    document.build(story)
    return output.getvalue()


def _entry_row(record: AccessRecord) -> list[str]:
    if record.student:
        return [record.datetime.strftime("%d/%m/%Y"), record.datetime.strftime("%H:%M:%S"),
                record.student.matricula, record.student.nombre_completo, "ALUMNO",
                record.student.carrera, record.student.grupo, record.student.plantel]
    if record.personnel:
        return [record.datetime.strftime("%d/%m/%Y"), record.datetime.strftime("%H:%M:%S"),
                record.personnel.numero_empleado, record.personnel.nombre_completo,
                record.personnel.tipo_personal, record.personnel.area or "—", "PERSONAL",
                record.personnel.plantel]
    return [record.datetime.strftime("%d/%m/%Y"), record.datetime.strftime("%H:%M:%S"),
            "—", "No identificado", "NO IDENTIFICADO", "—", "—", "—"]


def summarize_entries(records: list[AccessRecord]) -> dict:
    identities = set()
    students = set()
    personnel = set()
    careers: Counter[str] = Counter()
    for record in records:
        if record.student:
            identities.add(("student", record.student.id))
            students.add(record.student.id)
            careers[record.student.carrera or "Sin carrera"] += 1
        elif record.personnel:
            identities.add(("personnel", record.personnel.id))
            personnel.add(record.personnel.id)
    return {"entries": len(records), "unique_people": len(identities), "students": len(students),
            "personnel": len(personnel), "career_count": len(careers),
            "career_entries": sorted(careers.items(), key=lambda item: (-item[1], item[0]))}


def create_entries_csv(records: list[AccessRecord]) -> bytes:
    summary = summarize_entries(records)
    stream = io.StringIO(newline="")
    stream.write("\ufeff")
    writer = csv.writer(stream)
    writer.writerow(["REPORTE DE ENTRADAS"])
    writer.writerow(["Entradas", summary["entries"], "Personas únicas", summary["unique_people"],
                     "Alumnos", summary["students"], "Personal", summary["personnel"]])
    writer.writerow([])
    writer.writerow(ENTRY_HEADERS)
    writer.writerows(_entry_row(record) for record in records)
    return stream.getvalue().encode("utf-8")


def create_entries_excel(records: list[AccessRecord]) -> bytes:
    summary = summarize_entries(records)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Resumen"
    overview.append([settings.INSTITUTION_NAME])
    overview.append(["Reporte de entradas autorizadas"])
    overview.append(["Generado", datetime.now().strftime("%d/%m/%Y %H:%M")])
    overview.append([])
    overview.append(["Indicador", "Cantidad"])
    overview.append(["Eventos de entrada", summary["entries"]])
    overview.append(["Personas únicas", summary["unique_people"]])
    overview.append(["Alumnos únicos", summary["students"]])
    overview.append(["Personal único", summary["personnel"]])
    overview.append([])
    overview.append(["Carrera", "Entradas"])
    for career, count in summary["career_entries"]:
        overview.append([career, count])
    overview.column_dimensions["A"].width = 38
    overview.column_dimensions["B"].width = 18
    for cell in overview[5]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="173C35")
    detail = workbook.create_sheet("Quiénes entraron")
    detail.append(ENTRY_HEADERS)
    for record in records:
        detail.append(_entry_row(record))
    for cell in detail[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="173C35")
    for index, width in enumerate([13, 12, 18, 36, 22, 32, 18, 24], 1):
        detail.column_dimensions[chr(64 + index)].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    people_sheet = workbook.create_sheet("Personas únicas")
    people_headers = ["Identificador", "Persona", "Tipo", "Carrera o área", "Grupo", "Plantel",
                      "Primera entrada", "Última entrada", "Total entradas"]
    people_sheet.append(people_headers)
    grouped_people: dict[tuple[str, int], dict] = {}
    for record in records:
        if record.student:
            key = ("student", record.student.id)
            base = [record.student.matricula, record.student.nombre_completo, "ALUMNO",
                    record.student.carrera, record.student.grupo, record.student.plantel]
        elif record.personnel:
            key = ("personnel", record.personnel.id)
            base = [record.personnel.numero_empleado, record.personnel.nombre_completo,
                    record.personnel.tipo_personal, record.personnel.area or "—", "PERSONAL",
                    record.personnel.plantel]
        else:
            key = ("unknown", record.id)
            base = ["—", "No identificado", "NO IDENTIFICADO", "—", "—", "—"]
        item = grouped_people.setdefault(key, {"base": base, "times": []})
        item["times"].append(record.datetime)
    for item in sorted(grouped_people.values(), key=lambda value: value["base"][1]):
        times = sorted(item["times"])
        people_sheet.append(item["base"] + [times[0].strftime("%d/%m/%Y %H:%M"),
                                             times[-1].strftime("%d/%m/%Y %H:%M"), len(times)])
    for cell in people_sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="173C35")
    for index, width in enumerate([18, 36, 22, 32, 18, 24, 22, 22, 16], 1):
        people_sheet.column_dimensions[chr(64 + index)].width = width
    people_sheet.freeze_panes = "A2"
    people_sheet.auto_filter.ref = people_sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_entries_pdf(records: list[AccessRecord]) -> bytes:
    summary = summarize_entries(records)
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=10 * mm,
                                 leftMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph(settings.INSTITUTION_NAME, styles["Title"]),
             Paragraph("Reporte de entradas autorizadas", styles["Heading2"]),
             Paragraph(f"Entradas: {summary['entries']} · Personas únicas: {summary['unique_people']} · "
                       f"Alumnos: {summary['students']} · Personal: {summary['personnel']}", styles["BodyText"]),
             Spacer(1, 5 * mm)]
    career_data = [["Carrera", "Entradas"]] + [[name, count] for name, count in summary["career_entries"]]
    if len(career_data) > 1:
        career_table = Table(career_data, repeatRows=1, colWidths=[80 * mm, 25 * mm])
        career_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173C35")),
                                          ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                                          ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#D5DDD9")),
                                          ("FONTSIZE", (0, 0), (-1, -1), 8)]))
        story.extend([career_table, Spacer(1, 5 * mm)])
    data = [ENTRY_HEADERS] + [_entry_row(record) for record in records]
    table = Table(data, repeatRows=1, colWidths=[20*mm, 17*mm, 25*mm, 43*mm, 25*mm, 39*mm, 23*mm, 30*mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173C35")),
                               ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                               ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                               ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                               ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#D5DDD9")),
                               ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F5")])]))
    story.append(table)
    document.build(story)
    return output.getvalue()
