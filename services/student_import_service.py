from __future__ import annotations

import csv
import io
import secrets
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from config import settings
from database.models import Student
from services.qr_service import generate_student_qr


HEADERS = ["matricula", "nombres", "apellido_paterno", "apellido_materno", "carrera", "grupo",
           "plantel", "turno", "fecha_vencimiento", "activo"]
REQUIRED_HEADERS = ["matricula", "nombres", "apellido_paterno", "apellido_materno", "carrera",
                    "plantel", "turno", "fecha_vencimiento"]
VALID_SHIFTS = {"Matutino", "Vespertino", "Mixto", "Nocturno"}


def suggest_next_matricula(db: Session) -> str:
    prefix = str(date.today().year)
    existing = db.scalars(select(Student.matricula).where(
        Student.matricula.like(f"{prefix}%")).order_by(Student.matricula.desc()).limit(100)).all()
    numbers = [int(value[len(prefix):]) for value in existing if value[len(prefix):].isdigit()]
    return f"{prefix}{(max(numbers, default=0) + 1):04d}"


def sample_csv() -> bytes:
    stream = io.StringIO(newline="")
    stream.write("\ufeff")
    writer = csv.writer(stream)
    writer.writerow(HEADERS)
    writer.writerow([suggested_demo_matricula(), "María", "Ejemplo", "López", "Ingeniería en Sistemas",
                     "3A", "Campus Centro", "Matutino",
                     (date.today() + timedelta(days=365)).isoformat(), "SI"])
    return stream.getvalue().encode("utf-8")


def suggested_demo_matricula() -> str:
    return f"{date.today().year}9999"


def _read_rows(content: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        return [{str(k).strip(): v for k, v in row.items()} for row in csv.DictReader(io.StringIO(text))]
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:] if any(value not in (None, "") for value in row)]
    raise ValueError("Usa un archivo CSV o XLSX")


def _text(value, length: int) -> str:
    return " ".join(str(value or "").strip().split())[:length]


def _date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value or "").strip())


def _active(value) -> bool:
    return str(value or "SI").strip().upper() in {"SI", "SÍ", "1", "TRUE", "ACTIVO", "YES"}


def process_student_file(content: bytes, filename: str, db: Session, commit: bool = False) -> dict:
    if len(content) > 3_000_000:
        raise ValueError("El archivo excede 3 MB")
    rows = _read_rows(content, filename)
    if len(rows) > 500:
        raise ValueError("Importa como máximo 500 alumnos por archivo")
    if not rows:
        raise ValueError("El archivo no contiene alumnos")
    missing_headers = [header for header in REQUIRED_HEADERS if header not in rows[0]]
    if missing_headers:
        raise ValueError("Faltan columnas: " + ", ".join(missing_headers))

    seen: set[str] = set()
    valid: list[dict] = []
    errors: list[dict] = []
    for line, row in enumerate(rows, 2):
        matricula = _text(row.get("matricula"), 32)
        try:
            if not matricula:
                raise ValueError("Matrícula vacía")
            if matricula in seen:
                raise ValueError("Matrícula repetida en el archivo")
            if db.scalar(select(Student.id).where(Student.matricula == matricula)):
                raise ValueError("Matrícula ya registrada")
            seen.add(matricula)
            nombres = _text(row.get("nombres"), 120)
            paterno = _text(row.get("apellido_paterno"), 80)
            carrera = _text(row.get("carrera"), 160)
            plantel = _text(row.get("plantel"), 120)
            turno = _text(row.get("turno"), 40).title()
            if not all((nombres, paterno, carrera, plantel)):
                raise ValueError("Faltan datos obligatorios")
            if turno not in VALID_SHIFTS:
                raise ValueError("Turno no válido")
            expiry = _date(row.get("fecha_vencimiento"))
            if expiry < date.today():
                raise ValueError("La vigencia está vencida")
            valid.append({"matricula": matricula, "nombres": nombres, "apellido_paterno": paterno,
                          "apellido_materno": _text(row.get("apellido_materno"), 80),
                          "carrera": carrera, "grupo": _text(row.get("grupo"), 60) or "Sin grupo",
                          "plantel": plantel, "turno": turno,
                          "fecha_vencimiento": expiry, "activo": _active(row.get("activo"))})
        except (ValueError, TypeError) as exc:
            errors.append({"line": line, "matricula": matricula or "—", "error": str(exc)})

    created: list[Student] = []
    if commit and valid:
        for values in valid:
            student = Student(student_id=f"SAU-{secrets.token_hex(8)}", **values)
            db.add(student)
            created.append(student)
        db.commit()
        for student in created:
            db.refresh(student)
            generate_student_qr(student.id, db)
    return {"total": len(rows), "valid": len(valid), "errors": errors,
            "created": len(created), "committed": commit, "preview": valid[:15]}
